import os
from tkinter import *
import openpyxl as oxl
from tkinter import ttk
import customtkinter as ctk
from tabulate import tabulate
from PIL import Image ,ImageTk
from tkinter.messagebox import showerror

def Imgo(file,w,h) :

    # Image processing
    img=Image.open(file)
    pht=ImageTk.PhotoImage(img.resize((w,h), Image.Resampling.LANCZOS ))
    return pht

def change(can,page) :

    # Switching canvas
    can.destroy()
    page()

def hel() :

    # Pop up window
    showerror(title="Info",message="INVALID ENTRY")

def not_allow() :

    # Pop up window
    showerror(title="Info",message="SLOT FULL")

def compare_rec(app_no,pw_dob) :

    # Comparing ID and password
    global rec
    wb=oxl.load_workbook( os.path.join( os.getcwd(), "Data_Files\\student_data.xlsx") )
    ws=wb["Sheet1"]
    for i in range(2,ws.max_row+1) :
        flag=0
        x=ws[f"B{i}"]
        y=ws[f"L{i}"]
        if app_no == str(x.value) and pw_dob == y.value :
            flag=1
            rec=i
            break
    if flag==1 :
        change(frpa, secondpage)

    else :
        hel()

def colfee(data,x) :

    # Table of collage fee data
    s,f,c=[],[],[]
    for i in range(2,x) :
        a=data[i].split("-")
        s.append(a[0])
        f.append(a[1])
        c.append(a[2])
    dic={s[0]:s[1:]}
    c1=tabulate(dic,headers="keys",tablefmt="flat",colalign=['left'])
    dic={f[0]:f[1:]}
    c2=tabulate(dic,headers="keys",tablefmt="flat",colalign=['left'])
    dic={c[0]:c[1:]}
    c3=tabulate(dic,headers="keys",tablefmt="flat",colalign=['left'])
    return c1,c2,c3

def hostfee(data) :

    # Table of hostel fee data
    hn,hc,mc,sc,to=[],[],[],[],[]
    for i in range(14,19) :
        a=data[i].split("-")
        hn.append(a[0])
        hc.append(a[1])
        mc.append(a[2])
        sc.append(a[3])
        to.append(a[4])
    dic={hn[0]:hn[1:]}
    c1=tabulate(dic,headers="keys",tablefmt="flat",colalign=['left'])
    dic={hc[0]:hc[1:]}
    c2=tabulate(dic,headers="keys",tablefmt="flat",colalign=['left'])
    dic={mc[0]:mc[1:]}
    c3=tabulate(dic,headers="keys",tablefmt="flat",colalign=['left'])
    dic={sc[0]:sc[1:]}
    c4=tabulate(dic,headers="keys",tablefmt="flat",colalign=['left'])
    dic={to[0]:to[1:]}
    c5=tabulate(dic,headers="keys",tablefmt="flat",colalign=['left'])
    return c1,c2,c3,c4,c5

def propage() :

    # Managing data
    wb=oxl.load_workbook( os.path.join( os.getcwd(), "Data_Files\\student_data.xlsx") )
    ws=wb["Sheet1"]

    propa=Canvas(root,width=wid,height=hgt,borderwidth=0)
    propa.pack(fill="both",expand=True)

    # Background image
    back_img=Imgo( os.path.join( os.getcwd(), "Images\\front2c.jpg"),1700,810)
    propa.create_image(0,0,image=back_img,anchor="nw")

    # Back button
    back_bt=ctk.CTkButton(master=root, text="BACK", text_font=("Helvetica",20),  width=20, height=10,  corner_radius=10,
                        fg_color="red", hover_color="#ff5359", bg_color="#fafafa", text_color="#f9f9f9", border_width=0,
                        command= lambda : change(propa,secondpage))
    back_bt_win=propa.create_window(20,15,anchor="nw",window=back_bt)

    # Heading
    propa.create_text(750,50,text="Profile",font=("Book Antiqua",30,"bold","underline"),anchor="nw",fill="#ff512a")

    # Roll no
    propa.create_text(300,150,text="Roll No",font=("Book Antiqua",22,"bold"),anchor="nw",fill="black")
    propa.create_text(590,150,text=":",font=("Book Antiqua",22,"bold"),anchor="nw",fill="black")
    propa.create_text(700,150,text=ws[f"B{rec}"].value,font=("Book Antiqua",20,"bold"),anchor="nw",fill="white")

    # JEE Rank
    propa.create_text(300,200,text="JEE Rank",font=("Book Antiqua",22,"bold"),anchor="nw",fill="black")
    propa.create_text(590,200,text=":",font=("Book Antiqua",22,"bold"),anchor="nw",fill="black")
    propa.create_text(700,200,text=ws[f"H{rec}"].value,font=("Book Antiqua",20,"bold"),anchor="nw",fill="white")

    # Name
    propa.create_text(300,250,text="Name",font=("Book Antiqua",22,"bold"),anchor="nw",fill="black")
    propa.create_text(590,250,text=":",font=("Book Antiqua",22,"bold"),anchor="nw",fill="black")
    propa.create_text(700,250,text=ws[f"C{rec}"].value,font=("Book Antiqua",20,"bold"),anchor="nw",fill="white")

    # Father Name
    propa.create_text(300,300,text="Father Name",font=("Book Antiqua",22,"bold"),anchor="nw",fill="black")
    propa.create_text(590,300,text=":",font=("Book Antiqua",22,"bold"),anchor="nw",fill="black")
    propa.create_text(700,300,text=ws[f"D{rec}"].value,font=("Book Antiqua",20,"bold"),anchor="nw",fill="white")

    # Mobile no
    propa.create_text(300,350,text="Mobile No",font=("Book Antiqua",22,"bold"),anchor="nw",fill="black")
    propa.create_text(590,350,text=":",font=("Book Antiqua",22,"bold"),anchor="nw",fill="black")
    propa.create_text(700,350,text=ws[f"G{rec}"].value,font=("Book Antiqua",20,"bold"),anchor="nw",fill="white")

    # DOB
    propa.create_text(300,400,text="Date Of Birth",font=("Book Antiqua",22,"bold"),anchor="nw",fill="black")
    propa.create_text(590,400,text=":",font=("Book Antiqua",22,"bold"),anchor="nw",fill="black")
    propa.create_text(700,400,text=ws[f"L{rec}"].value,font=("Book Antiqua",20,"bold"),anchor="nw",fill="white")

    # Gender
    propa.create_text(300,450,text="Gender",font=("Book Antiqua",22,"bold"),anchor="nw",fill="black")
    propa.create_text(590,450,text=":",font=("Book Antiqua",22,"bold"),anchor="nw",fill="black")
    propa.create_text(700,450,text=ws[f"F{rec}"].value,font=("Book Antiqua",20,"bold"),anchor="nw",fill="white")

    # 10th marks
    propa.create_text(300,500,text="10th Marks",font=("Book Antiqua",22,"bold"),anchor="nw",fill="black")
    propa.create_text(590,500,text=":",font=("Book Antiqua",22,"bold"),anchor="nw",fill="black")
    propa.create_text(700,500,text=ws[f"I{rec}"].value,font=("Book Antiqua",20,"bold"),anchor="nw",fill="white")

    # 12th marks
    propa.create_text(300,550,text="12th Marks",font=("Book Antiqua",22,"bold"),anchor="nw",fill="black")
    propa.create_text(590,550,text=":",font=("Book Antiqua",22,"bold"),anchor="nw",fill="black")
    propa.create_text(700,550,text=ws[f"J{rec}"].value,font=("Book Antiqua",20,"bold"),anchor="nw",fill="white")

    # Branch
    propa.create_text(300,600,text="Branch",font=("Book Antiqua",22,"bold"),anchor="nw",fill="black")
    propa.create_text(590,600,text=":",font=("Book Antiqua",22,"bold"),anchor="nw",fill="black")
    propa.create_text(700,600,text=ws[f"E{rec}"].value,font=("Book Antiqua",20,"bold"),anchor="nw",fill="white")

    # Mail ID
    propa.create_text(300,650,text="Mail ID",font=("Book Antiqua",22,"bold"),anchor="nw",fill="black")
    propa.create_text(590,650,text=":",font=("Book Antiqua",22,"bold"),anchor="nw",fill="black")
    propa.create_text(700,650,text=ws[f"K{rec}"].value,font=("Book Antiqua",20,"bold"),anchor="nw",fill="white")

    root.mainloop()

def slotpage() :

    # Defining variables
    global date_drop
    global slot_drop
    global save_bt
    global slotp
    global slotpa
    global choose_date
    global time1
    global time2

    # Managing file
    wb=oxl.load_workbook("Data_Files\\student_data.xlsx")
    ws2=wb["Sheet2"]
    ws3=wb["Sheet3"]

    slotpa = Frame(master=root)
    slotp=Canvas( slotpa, width=wid, height=hgt, borderwidth=0)
    slotp.pack(fill="both",expand=True)

    # Background image
    back_img=Imgo( os.path.join( os.getcwd(), "Background\\front2d.jpg"),1700,810)
    slotp.create_image(0,0,image=back_img,anchor="nw")

    # Heading
    slotp.create_text(700,65, text="Slot Booking", font=("Book Antiqua",28,"bold","underline"), anchor="nw", fill="#ff512a")

    # Round info
    slotp.create_text(220+100,200, text="Round Alloted :", font=("Book Antiqua",22,"bold"), anchor="nw", fill="#000000")
    slotp.create_text(620+420,200, text=ws2[f"B{rec}"].value, font=("Book Antiqua",22,"bold"), anchor="nw", fill="#f9f9f9")

    # Date label
    slotp.create_text(220+100,275, text="Choose A Date out of following -", font=("Book Antiqua",22,"bold","underline"), anchor="nw", fill="black")
    slotp.create_text(220+100,350, text="Select Date :", font=("Book Antiqua",22,"bold"), anchor="nw", fill="black")

    # Slot label
    slotp.create_text(220+100,425, text="Choose a time slot for Counselling -", font=("Book Antiqua",22,"bold","underline"), anchor="nw", fill="black")
    slotp.create_text(220+100,500, text="Select Time Slot :", font=("Book Antiqua",22,"bold"), anchor="nw", fill="black")

    choose_date , time1 , time2 = limit_check()

    # Back button
    back_bt=ctk.CTkButton(master=slotpa, text="BACK", text_font=("Helvetica",20),  width=20, height=10, corner_radius=10,
                        fg_color="red", hover_color="#ff5359", bg_color="#ebebeb", text_color="#f9f9f9", border_width=0,
                        command= lambda : change_slot(slotp,secondpage,slotpa))
    back_bt_win=slotp.create_window(20,15,anchor="nw",window=back_bt)

    root.mainloop()

def docpage() :

    # Defining Structure
    docpa=Canvas(root,width=wid,height=hgt,borderwidth=0)
    docpa.pack(fill="both",expand=True)

    # Background Image
    back_img=Imgo( os.path.join( os.getcwd(), "Background\\front2e.jpg"),1700,810)
    docpa.create_image(0,0,image=back_img,anchor="nw")

    # Heading
    docpa.create_text(450,65,text="Documents Require For Counselling",font=("Book Antiqua",30,"bold","underline"),
                        anchor="nw",fill="#ff512a")

    # File handling
    with open( os.path.join( os.getcwd(), "Data_Files\\coun_doc.txt") ) as myfile:
        fees_info=myfile.readlines()

    # Table Structure and data
    c1,c2,c3=colfee(fees_info,15)
    docpa.create_text(250,160, text=c1, font=("Book Antiqua",20,"bold"), anchor="nw", fill="#f9f9f9")
    docpa.create_text(430,160, text=c2, font=("Book Antiqua",20,"bold"), anchor="nw", fill="#f9f9f9")
    docpa.create_text(1210,160,text=c3, font=("Book Antiqua",20,"bold"), anchor="nw", fill="#f9f9f9")

    # Back button
    back_bt=ctk.CTkButton(master=root, text="BACK", text_font=("Helvetica",20),  width=20, height=10,corner_radius=10,
                        fg_color="red", hover_color="#ff5359", bg_color="#eaeaea", text_color="#f9f9f9", border_width=0,
                        command= lambda : change(docpa,secondpage))
    back_bt_win=docpa.create_window(20,15,anchor="nw",window=back_bt)

    root.mainloop()

def collpage() :

    # File handling
    with open( os.path.join( os.getcwd(), "Data_Files\\fees_detail.txt") ) as myfile:
        fees_info=myfile.readlines()

    # Background Image
    back_img=Imgo( os.path.join( os.getcwd(), "Background\\front2f.jpg"),1700,810)
    collpa.create_image(0,0,image=back_img,anchor="nw")

    # Heading
    collpa.create_text(480,65,text="JSS FEE OF I YEAR, 2022-23 BATCH",
                       font=("Book Antiqua",28,"bold","underline"), anchor="nw", fill="#ff512a")
    c1,c2,c3=colfee(fees_info,11)

    # Table Structure
    collpa.create_text(250,160,text=c1,font=("Book Antiqua",20,"bold"),anchor="nw",fill="white")
    collpa.create_text(450,160,text=c2,font=("Book Antiqua",20,"bold"),anchor="nw",fill="white")
    collpa.create_text(1200,160,text=c3,font=("Book Antiqua",20,"bold"),anchor="nw",fill="white")

    #Payment Mode
    collpa.create_text(250,580,text="MODE OF PAYMENT -",font=("Book Antiqua",22,"bold","underline"),anchor="nw",fill="white")
    collpa.create_text(280,640,text="Demand Draft in favour of ",font=("Book Antiqua",18),anchor="nw",fill="white")
    collpa.create_text(645,640,text="*JSSATE MANAGEMENT A/C*",font=("Book Antiqua",18,"bold"),anchor="nw",fill="white")
    collpa.create_text(1095,640,text="Payable at Noida / New Delhi",font=("Book Antiqua",18),anchor="nw",fill="white")

    #Account No
    collpa.create_text(280,690,text="Account No. -",font=("Book Antiqua",18,"bold","underline"),anchor="nw",fill="white")
    collpa.create_text(480,690,text=" 520101236640492",font=("Book Antiqua",18),anchor="nw",fill="white")

    #IFSC Code
    collpa.create_text(930,690,text="IFSC Code -",font=("Book Antiqua",18,"bold","underline"),anchor="nw",fill="white")
    collpa.create_text(1110,690,text=" UBIN0920797",font=("Book Antiqua",18),anchor="nw",fill="white")

    #Bank Name
    collpa.create_text(280,740,text="Bank Name -",font=("Book Antiqua",18,"bold","underline"),anchor="nw",fill="white")
    collpa.create_text(480,740,text=" UNION BANK OF INDIA",font=("Book Antiqua",18),anchor="nw",fill="white")

    # Back button
    back_bt=ctk.CTkButton(master=root, text="BACK", text_font=("Helvetica",20), width=20, height=10, corner_radius=10,
                         fg_color="red", hover_color="#ff5359", bg_color="#fafafa", text_color="#f9f9f9", border_width=0,
                         command= lambda : change(collpa,secondpage))
    back_bt_win=collpa.create_window(20,15,anchor="nw",window=back_bt)

    root.mainloop()

def hostpage() :

    hostpa=Canvas(root,width=wid,height=hgt,borderwidth=0)
    hostpa.pack(fill="both",expand=True)

    # Background image
    back_img=Imgo( os.path.join( os.getcwd(), "Background\\front2g.jpg"),1700,810)
    hostpa.create_image(0,0,image=back_img,anchor="nw")

    # Heading of page
    hostpa.create_text(370,80,text="JSS HOSTEL FEE OF I YEAR, 2022-23 BATCH",font=("Book Antiqua",28,"bold","underline"),
                       anchor="nw",fill="#ff512a")

    # File handlling
    with open( os.path.join( os.getcwd(), "Data_Files\\fees_detail.txt") ) as myfile:
        fees_info=myfile.readlines()
    c1,c2,c3,c4,c5=hostfee(fees_info)
    
    # Structure of data
    hostpa.create_text(200,220,text=c1,font=("Book Antiqua",22,"bold"),anchor="nw",fill="white")
    hostpa.create_text(500,220,text=c2,font=("Book Antiqua",22,"bold"),anchor="nw",fill="white")
    hostpa.create_text(800,220,text=c3,font=("Book Antiqua",22,"bold"),anchor="nw",fill="white")
    hostpa.create_text(1080,220,text=c4,font=("Book Antiqua",22,"bold"),anchor="nw",fill="white")
    hostpa.create_text(1280,220,text=c5,font=("Book Antiqua",22,"bold"),anchor="nw",fill="white")
    
    # Defining details
    hostpa.create_text(200,525,text="MODE OF PAYMENT : Only Through Demand Draft",font=("Book Antiqua",22,"bold","underline"),
                        anchor="nw",fill="white")
    hostpa.create_text(230,580,text="Demand Draft in favour of *JSSATE - Hostel Establishment A/C*",font=("Book Antiqua",22),
                        anchor="nw",fill="white")
    hostpa.create_text(230,620,text="Payable at Noida / New Delhi",font=("Book Antiqua",22),anchor="nw",fill="white")

    # Back button
    back_bt=ctk.CTkButton(master=root, text="BACK", text_font=("Helvetica",20),  width=20, height=10, corner_radius=10,
                      fg_color="red", hover_color="#ff5359", bg_color="#ebebeb",text_color="#f9f9f9", border_width=0,
                      command= lambda : change(hostpa,secondpage))
    back_bt_win=hostpa.create_window(20,15,anchor="nw",window=back_bt)

    root.mainloop()

def secondpage():

    global scpa

    # Data managing
    wb=oxl.load_workbook( os.path.join( os.getcwd(), "Data_Files\\student_data.xlsx") )
    ws=wb["Sheet1"]
    scpa=Canvas(root,width=1560,height=810,borderwidth=0)
    scpa.pack(fill="both",expand=True)

    # JSS logo image
    jss_image=Imgo( os.path.join( os.getcwd(), "Images\\jss.png"),160,140)

    # Background Image
    back2_image=Imgo( os.path.join( os.getcwd(), "Background\\front2b.jpg"),1700,810)
    scpa.create_image(0,0,image=back2_image,anchor="nw")

    # Welocome Text
    scpa.create_text( 220, 100-10, text="Welcome ,", font=("Book Antiqua",26,"bold"), anchor="nw", fill="#ff512a")
    scpa.create_text( 460, 80-10, text=ws[f"C{rec}"].value, font=("Book Antiqua",38,"bold"), anchor="nw", fill="#ff512a")

    # Profile page window
    pro=Imgo( os.path.join( os.getcwd(), "Images\\profile.jpg"),220,170)
    pro_bt=ctk.CTkButton(master=root,image=pro, text="Profile", text_font=("Book Antiqua",22,"bold"), compound="top",
                         corner_radius=10, bg_color="#fafafa", fg_color="#2d435b", hover_color="#fdbf38", text_color="white",
                         width=230, height=200, border_width=0, command= lambda : change(scpa,propage))
    pro_bt_win=scpa.create_window(420,200-10,anchor="nw",window=pro_bt)

    # Slot page window
    slot=Imgo( os.path.join( os.getcwd(), "Images\\clo.png"),220,170)
    slot_bt=ctk.CTkButton(master=root,image=slot, text="Slot Booking", text_font=("Book Antiqua",22,"bold"), compound="top",
                         corner_radius=10, bg_color="#fafafa", fg_color="#2d435b", hover_color="#fdbf38", text_color="white",
                         width=230, height=200, border_width=0, command= lambda : change(scpa,slotpage))
    slot_bt_win=scpa.create_window(740,200-10,anchor="nw",window=slot_bt)

    # Document page window
    doc=Imgo( os.path.join( os.getcwd(), "Images\\doc_img2.png"),220,170)
    doc_bt=ctk.CTkButton(master=root,image=doc, text="Documents", text_font=("Book Antiqua",22,"bold"), compound="top",
                         corner_radius=10, bg_color="#fafafa", fg_color="#2d435b", hover_color="#fdbf38", text_color="white",
                         width=230, height=200, border_width=0, command= lambda : change(scpa,docpage))
    doc_bt_win=scpa.create_window(1060,200-10,anchor="nw",window=doc_bt)

    # College page window
    coll=Imgo( os.path.join( os.getcwd(), "Images\\college.jpg"),220,170)
    coll_bt=ctk.CTkButton(master=root,image=coll, text="College Fees", text_font=("Book Antiqua",22,"bold"), compound="top",
                         corner_radius=10, bg_color="#fafafa", fg_color="#2d435b", hover_color="#fdbf38", text_color="white",
                         width=230, height=200, border_width=0, command=lambda : change(scpa,collpage))
    coll_bt_win=scpa.create_window(550,475-10,anchor="nw",window=coll_bt)

    # Hostel page window
    host=Imgo( os.path.join( os.getcwd(), "Images\\hostel.jpg"),220,170)
    host_bt=ctk.CTkButton(master=root,image=host, text="Hostel Fees", text_font=("Book Antiqua",22,"bold"), compound="top",
                         corner_radius=10, bg_color="#fafafa", fg_color="#2d435b", hover_color="#fdbf38", text_color="white",
                         width=230, height=200, border_width=0, command= lambda : change(scpa,hostpage))
    host_bt_win=scpa.create_window(870,475-10,anchor="nw",window=host_bt)

    # Logout button
    logout_bt=ctk.CTkButton(master=root, text="Log Out", text_font=("Book Antiqua",22,"bold"),  width=30, height=20, corner_radius=10,
                         fg_color="red", hover_color="#ff5359", bg_color="#fafafa",border_width=0,
                         command=lambda : change(scpa,firstpage))
    logout_bt_win=scpa.create_window(760,750-10,anchor="nw",window=logout_bt)

    root.mainloop()

def firstpage() :

    global user,pwrd,frpa

    # Defining Structure
    frpa=Canvas( root, width = wid, height = hgt, bg = "black", highlightcolor = "#3c5390", borderwidth = 0 )
    frpa.pack( fill = "both", expand = True )

    # Image on top
    jss_image = Imgo(os.path.join( os.getcwd(), "Images\\jss.png"), 135, 135)
    # Background Image
    back_image = Imgo(os.path.join( os.getcwd(), "Background\\front2a.jpg"), 1700, 810)
    entry_image = Imgo(os.path.join( os.getcwd(), "Images\\front22.png"), 550, 370)
    frpa.create_image( 0, 0, image = back_image , anchor = "nw")
    frpa.create_image(30, 20, image = jss_image, anchor = "nw")
    frpa.create_image(750, 150, image = entry_image, anchor = "nw")
    frpa.create_image(750, 350, image = entry_image, anchor = "nw")

    # Heading
    frpa.create_text(915,80,text="JSS ACADEMY OF TECHNICAL EDUCATION, NOIDA",font=("Book Antiqua",34,"bold"),fill="red")
    frpa.create_text(1025,180,text="JSS ADMISSION",font=("Book Antiqua",28,"bold"),fill="#0b4bf5")##0b4bf5
    frpa.create_text(970,305,text="Username",font=("Book Antiqua",22,"bold"),fill="white")
    frpa.create_text(970,505,text="Password",font=("Book Antiqua",22,"bold"),fill="white")
    
    # Entry of username and password
    user = ctk.CTkEntry( master = frpa, 
                          placeholder_text = "Roll Number", text_font = ( "Book Antiqua", 20, "bold" ), 
                           width = 230, height = 30, corner_radius = 10,
                            placeholder_text_color = "white", text_color = "white", 
                             fg_color = "#ffca62", bg_color = "#ffa900", 
                              border_color = "#162d50", border_width = 0)
    user_win = frpa.create_window( 880, 355, anchor = "nw", window = user )

    pwrd = ctk.CTkEntry( master = frpa, 
                          placeholder_text = "DOB(dd/mm/yyyy)", text_font = ( "Book Antiqua", 20, "bold" ), 
                           width = 250, height = 30, corner_radius = 10,
                            placeholder_text_color = "white", text_color = "white", 
                             fg_color = "#ffca62", bg_color = "#ffa900", 
                              border_color = "#162d50", border_width = 0, show = "*" )
    pwrd_win = frpa.create_window( 875, 555, anchor = "nw", window = pwrd )

    # Login button
    log_bt=ctk.CTkButton(master=root, text="Login", text_font=("Book Antiqua",25,"bold"), width=30, height=20,
                         corner_radius=15, text_color="white", bg_color="#dadada", fg_color="#162d50", 
                         hover_color="#0c3f8c", border_width=0, command = lambda : compare_rec( user.get(), pwrd.get()) )
    log_bt_win=frpa.create_window(960,675,anchor="nw",window=log_bt)

    root.mainloop()

if __name__ == "__main__" :

    # Defining Main theme of all widgets
    ctk.set_appearance_mode("dark")
    ctk.set_default_color_theme("dark-blue")
    wid = 1350
    hgt = 650

    global root
    root=ctk.CTk()
    root.title("JSS Counselling System")
    root.iconbitmap(os.path.join( os.getcwd(), "Images\\icon.ico"))
    root.geometry("1350x650+100+80")
    root.resizable(False,False)
    firstpage()
